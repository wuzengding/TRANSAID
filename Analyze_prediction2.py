import argparse
import pickle
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
import seaborn as sns
from collections import defaultdict
import os
import random
from Bio import SeqIO
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class TranscriptAnalyzer:
    def __init__(self, predictions_file, fasta_file, prefix):
        """初始化分析器"""
        # 读取预测结果
        with open(predictions_file, 'rb') as f:
            raw_data = pickle.load(f)
            self.data = {item['transcript_id']: {
                'predictions': item['predictions'],
                'true_labels': item['true_labels'],
                'length': item['length'],
                'is_match': item['is_match']
            } for item in raw_data}
            
        # 读取序列数据
        self.sequences = {}
        for record in SeqIO.parse(fasta_file, "fasta"):
            transcript_id = record.id.split('.')[0]
            self.sequences[transcript_id] = str(record.seq)
        self.formatted_sequence_stats = defaultdict(int)
        self.prefix = prefix
        
    def normalize_transcript(self, predictions, length):
        """将转录本标准化到0-1区间"""
        norm_positions = np.linspace(0, 1, length)
        return norm_positions, predictions
        
    def find_continuous_regions(self, predictions, label_type):
        """找到连续的预测区域"""
        regions = []
        current_start = None
        current_length = 0
        
        for i, pred in enumerate(predictions):
            if pred == label_type:
                if current_start is None:
                    current_start = i
                current_length += 1
            else:
                if current_length > 0:
                    regions.append({
                        'start': current_start,
                        'length': current_length,
                        'rel_start': current_start / len(predictions)
                    })
                current_start = None
                current_length = 0
                
        if current_length > 0:
            regions.append({
                'start': current_start,
                'length': current_length,
                'rel_start': current_start / len(predictions)
            })
            
        return regions

    def analyze_transcript(self, transcript_id):
        """分析单个转录本"""
        transcript_data = self.data[transcript_id]
        predictions = transcript_data['predictions']
        length = transcript_data['length']
        
        # 找到TIS和TTS区域
        tis_regions = self.find_continuous_regions(predictions, 0)
        tts_regions = self.find_continuous_regions(predictions, 1)
        
        return {
            'tis_regions': tis_regions,
            'tts_regions': tts_regions,
            'normalized_predictions': self.normalize_transcript(predictions, length)
        }

    def plot_transcript_heatmap(self, output_dir, group_size=400):
        """
        绘制按长度分组的转录本预测热图,每组固定group_size个转录本
        """
        # 获取所有转录本长度并排序
        transcript_lengths = [(tid, d['length']) for tid, d in self.data.items()]
        transcript_lengths.sort(key=lambda x: x[1])
        
        # 将转录本分组
        num_groups = (len(transcript_lengths) + group_size - 1) // group_size
        groups = [transcript_lengths[i*group_size : (i+1)*group_size] 
                 for i in range(num_groups)]
        
        # 创建颜色映射 (TIS=红色, TTS=蓝色, non-TIS/TTS=浅灰色, 背景=白色)
        colors = ['red', 'blue', '#EEEEEE', 'white']
        cmap = ListedColormap(colors)
        
        # 为每个组绘制热图
        for idx, group in enumerate(groups):
            group_min = group[0][1]
            group_max = group[-1][1]
            
            # 创建数据矩阵
            data_matrix = np.full((len(group), group_max), 3)  # 背景为白色
            
            # 填充预测数据
            for row, (tid, length) in enumerate(group):
                predictions = self.data[tid]['predictions']
                data_matrix[row, :length] = predictions
            
            # 绘制热图
            plt.figure(figsize=(15, 10))
            
            # 绘制热图并保存返回的图像对象
            im = plt.imshow(data_matrix, aspect='auto', cmap=cmap, interpolation='none')
            
            # 设置标题和标签
            plt.title(f'{self.prefix} Length {group_min}-{group_max}bp (n={len(group)})')
            plt.ylabel('Transcripts')
            plt.xlabel('Position (bp)')

            # 修改后的 colorbar 创建方式
            cbar = plt.colorbar(im, ticks=[0, 1, 2])
            cbar.ax.set_yticklabels(['TIS', 'TTS', 'non-TIS/TTS'])  # 设置标签
            plt.yticks([])
            
            # 保存图片
            output_file = os.path.join(output_dir, 
                         f"{self.prefix}_heatmap_length{group_min}-{group_max}.png")
            plt.savefig(output_file, dpi=300, bbox_inches='tight')
            plt.close()
            print(f"Saved heatmap to {output_file}")

    def plot_density_distribution(self, output_dir):
        """绘制TIS/TTS位置分布密度图"""
        plt.figure(figsize=(12, 6))
        
        # 收集TIS和TTS位置
        tis_positions = []
        tts_positions = []
        
        for transcript_id in self.data:
            analysis = self.analyze_transcript(transcript_id)
            tis_positions.extend([r['rel_start'] for r in analysis['tis_regions']])
            tts_positions.extend([r['rel_start'] for r in analysis['tts_regions']])
        
        # 绘制密度图
        sns.kdeplot(data=tis_positions, color='red', label='TIS')
        sns.kdeplot(data=tts_positions, color='blue', label='TTS')
        
        plt.xlabel('Relative Position in Transcript')
        plt.ylabel('Density')
        plt.title(f'{self.prefix} TIS/TTS Position Distribution')
        plt.legend()
        
        output_file = os.path.join(output_dir, f"{self.prefix}_position_density.png")
        plt.savefig(output_file)
        plt.close()
        print(f"Saved density plot to {output_file}")

    def plot_continuity_stats(self, output_dir):
        """绘制连续区域长度分布"""
        tis_lengths = []
        tts_lengths = []
        
        for transcript_id in self.data:
            analysis = self.analyze_transcript(transcript_id)
            tis_lengths.extend([r['length'] for r in analysis['tis_regions']])
            tts_lengths.extend([r['length'] for r in analysis['tts_regions']])
        
        plt.figure(figsize=(12, 6))
        plt.hist([tis_lengths, tts_lengths], 
                label=['TIS', 'TTS'], 
                bins=30, 
                alpha=0.7)
        plt.xlabel('Continuous Region Length (bp)')
        plt.ylabel('Frequency')
        plt.title(f'{self.prefix} Distribution of Continuous TIS/TTS Regions')
        plt.legend()
        
        output_file = os.path.join(output_dir, f"{self.prefix}_continuity_distribution.png")
        plt.savefig(output_file)
        plt.close()
        print(f"Saved continuity distribution to {output_file}")

    def generate_statistics(self):
        """生成统计摘要"""
        stats = {
            'total_transcripts': len(self.data),
            'matching_transcripts': sum(1 for d in self.data.values() if d['is_match']),
            'tis_stats': defaultdict(float),
            'tts_stats': defaultdict(float),
            'transcript_types': defaultdict(int)
        }
        
        # 统计每种转录本类型的数量
        for tid in self.data:
            trans_type = tid.split('_')[0]
            stats['transcript_types'][trans_type] += 1
        
        # TIS和TTS统计
        for transcript_id in self.data:
            analysis = self.analyze_transcript(transcript_id)
            
            # TIS统计
            tis_regions = analysis['tis_regions']
            stats['tis_stats']['total_regions'] += len(tis_regions)
            if tis_regions:
                stats['tis_stats']['avg_length'] += sum(r['length'] for r in tis_regions)
                stats['tis_stats']['max_length'] = max(
                    stats['tis_stats']['max_length'],
                    max(r['length'] for r in tis_regions)
                )
            
            # TTS统计
            tts_regions = analysis['tts_regions']
            stats['tts_stats']['total_regions'] += len(tts_regions)
            if tts_regions:
                stats['tts_stats']['avg_length'] += sum(r['length'] for r in tts_regions)
                stats['tts_stats']['max_length'] = max(
                    stats['tts_stats']['max_length'],
                    max(r['length'] for r in tts_regions)
                )
        
        # 计算平均值
        if stats['tis_stats']['total_regions'] > 0:
            stats['tis_stats']['avg_length'] /= stats['tis_stats']['total_regions']
        if stats['tts_stats']['total_regions'] > 0:
            stats['tts_stats']['avg_length'] /= stats['tts_stats']['total_regions']
            
        return stats
    
    def collect_pattern_statistics(self):
        """
        收集所有格式化序列的统计信息
        """
        pattern_stats = defaultdict(int)
        total_count = 0
        
        for transcript_id in self.data.keys():
            result = self.format_sequence_with_predictions(transcript_id)
            pattern = result['formatted_seq']
            pattern_stats[pattern] += 1
            total_count += 1
        
        # 转换为排序后的列表
        sorted_stats = [(pattern, count, count/total_count*100) 
                       for pattern, count in pattern_stats.items()]
        sorted_stats.sort(key=lambda x: x[1], reverse=True)
        
        return sorted_stats, total_count


    def find_complementary_sites(self, sequence, current_site_type, current_pos):
        """
        Find potential complementary translation sites.
        
        Args:
            sequence (str): The complete transcript sequence
            current_site_type (str): 'start' or 'stop' indicating what we're looking for
            current_pos (int): Position of the current site
            
        Returns:
            list: List of tuples containing (position, codon)
        """
        complementary_sites = []
        seq_len = len(sequence)
        
        if current_site_type == 'start':
            # Looking for stop codons downstream
            pos = current_pos + 3  # Start after ATG
            while pos + 2 < seq_len:
                codon = sequence[pos:pos+3].upper()
                if codon in ['TAA', 'TAG', 'TGA']:
                    complementary_sites.append((pos, codon))
                pos += 3  # Move to next codon
        else:
            # Looking for start codons upstream
            pos = current_pos - 3  # Start before stop codon
            while pos >= 0:
                codon = sequence[pos:pos+3].upper()
                if codon == 'ATG':
                    complementary_sites.append((pos, codon))
                pos -= 3  # Move to previous codon
                
        return complementary_sites
        
    def find_triplet_stats(self, start_pos, stop_pos):
        """
        Calculate triplet codon statistics between start and stop positions.
        
        Args:
            start_pos (int or None): Position of start codon
            stop_pos (int or None): Position of stop codon
            
        Returns:
            dict: Dictionary containing in_frame status, codon count, and remaining bases
        """
        # Handle cases where either position is None
        if start_pos is None or stop_pos is None:
            return {
                'in_frame': 'N/A',
                'codon_count': 0,
                'remaining_bases': 0
            }
            
        # Ensure stop_pos is greater than start_pos
        if stop_pos <= start_pos:
            return {
                'in_frame': 'N/A',
                'codon_count': 0,
                'remaining_bases': 0
            }
        
        distance = stop_pos - start_pos
        codon_count = distance // 3
        remaining_bases = distance % 3
        is_triplet = remaining_bases == 0
        
        return {
            'in_frame': 'Yes' if is_triplet else 'No',
            'codon_count': codon_count,
            'remaining_bases': remaining_bases
        }

    def evaluate_kozak_sequence(self, sequence, atg_pos):
        """
        Evaluate Kozak sequence strength based on position weight matrix.
        
        Args:
            sequence (str): Full transcript sequence
            atg_pos (int): Position of the ATG start codon
            
        Returns:
            tuple: (kozak_sequence, strength)
        """
        if atg_pos < 6 or atg_pos + 4 >= len(sequence):
            return None, None
            
        # Extract Kozak context (-6 to +4)
        kozak = sequence[atg_pos-6:atg_pos+4].upper()
        if len(kozak) != 10:
            return None, None
            
        # Check core positions
        minus_3 = kozak[3]  # -3 position
        plus_4 = kozak[9]   # +4 position
        
        # Determine strength
        if minus_3 in ['A', 'G'] and plus_4 == 'G':
            strength = 'S'  # Strong
        elif minus_3 in ['A', 'G'] or plus_4 == 'G':
            strength = 'M'  # Moderate
        else:
            return None, None  # Don't report weak sequences
            
        return kozak, strength

    def find_kozak_sequences(self, sequence):
        """
        Find all valid Kozak sequences in the transcript.
        
        Args:
            sequence (str): Full transcript sequence
            
        Returns:
            list: List of formatted Kozak sequences with positions and strengths
        """
        kozak_sites = []
        pos = 0
        
        while True:
            pos = sequence.upper().find('ATG', pos)
            if pos == -1:
                break
                
            kozak_seq, strength = self.evaluate_kozak_sequence(sequence, pos)
            if kozak_seq and strength:
                # Format: first 6 bases + ATG with position + last base + strength
                formatted_kozak = (
                    f"{kozak_seq[:6].lower()}"
                    f"ATG<sub>{pos+1}</sub>"
                    f"{kozak_seq[9].lower()}"
                    f"({strength})"
                )
                kozak_sites.append(formatted_kozak)
            
            pos += 3
            
        return '; '.join(kozak_sites) if kozak_sites else ""

    def format_pattern(self, predictions, sequence):
        """
        Create simplified pattern showing only TIS and TTS positions.
    
        Args:
            predictions (np.array): Prediction array
            sequence (str): Original sequence
        
        Returns:
            str: Formatted pattern string
        """
        pattern_parts = []
        i = 0
        current_type = None
        current_bases = []
    
        while i < len(predictions):
            if predictions[i] in [0, 1]:
                if current_type is None:
                    current_type = predictions[i]
                    current_bases = [sequence[i].upper()]
                elif current_type == predictions[i]:
                    current_bases.append(sequence[i].upper())
                else:
                    # Add current group
                    color = 'red' if current_type == 0 else 'blue'
                    pattern_parts.append(
                        f'<span style="color:{color}">{"".join(current_bases)}</span>'
                    )
                    current_type = predictions[i]
                    current_bases = [sequence[i].upper()]
                i += 1
            else:
                if current_bases:
                    color = 'red' if current_type == 0 else 'blue'
                    pattern_parts.append(
                        f'<span style="color:{color}">{"".join(current_bases)}</span>'
                    )
                    current_type = None
                    current_bases = []
                if len(pattern_parts) > 0 and pattern_parts[-1] != '<span style="color:gray">—</span>':
                    pattern_parts.append('<span style="color:gray">—</span>')
                while i < len(predictions) and predictions[i] not in [0, 1]:
                    i += 1
        
        if current_bases:
            color = 'red' if current_type == 0 else 'blue'
            pattern_parts.append(
                f'<span style="color:{color}">{"".join(current_bases)}</span>'
            )
        
        return ''.join(pattern_parts) if pattern_parts else '<span style="color:gray">—</span>'
    
    def generate_formatted_sequences_report(self, output_dir):
        """
        Generate enhanced formatted sequence report with new columns.
        """
        html_content = """
        <html>
        <head>
            <style>
                table { border-collapse: collapse; width: 100%; }
                th, td { padding: 8px; text-align: left; border: 1px solid #ddd; }
                th { background-color: #f2f2f2; }
                .sequence { font-family: monospace; }
            </style>
        </head>
        <body>
            <h2>Formatted Sequence Report</h2>
            <table>
                <tr>
                    <th>Transcript ID</th>
                    <th>Formatted Sequence</th>
                    <th>Formatted Sequence with Position</th>
                    <th>Formatted Pattern</th>
                    <th>Potential Complementary Sites</th>
                    <th>Potential Kozak</th>
                    <th colspan="3">Right Triplet?</th>
                </tr>
                <tr>
                    <th colspan="6"></th>
                    <th>In Frame?</th>
                    <th>Codon Count</th>
                    <th>Remaining Bases</th>
                </tr>

        """
        
        # Track statistics for patterns
        pattern_stats = defaultdict(lambda: {'total': 0, 'with_complementary': 0})
        
        # Process each transcript
        results = []  # Store results for Excel export
        for transcript_id in sorted(self.data.keys()):
            sequence = self.sequences[transcript_id]
            predictions = self.data[transcript_id]['predictions']
            
            result = self.format_sequence_with_predictions(transcript_id)
            pattern = self.format_pattern(predictions, sequence)
            kozak_sequences = self.find_kozak_sequences(sequence)
            
            # Calculate triplet statistics
            found_start = None
            found_stop = None
            for i, pred in enumerate(predictions):
                if pred == 0 and found_start is None:  # First TIS
                    found_start = i
                elif pred == 1 and found_stop is None:  # First TTS
                    found_stop = i
                    
            triplet_stats = {'in_frame': 'N/A', 'codon_count': 0, 'remaining_bases': 0}
            if found_start is not None and found_stop is not None:
                triplet_stats = self.find_triplet_stats(found_start, found_stop)
            
            # Update pattern statistics
            pattern_stats[pattern]['total'] += 1
            if result['complementary_sites']:
                pattern_stats[pattern]['with_complementary'] += 1
            
            # Store row for HTML and Excel
            row_data = {
                'transcript_id': result['transcript_id'],
                'formatted_seq': result['formatted_seq'],
                'formatted_pattern': pattern,
                'formatted_seq_with_pos': result['formatted_seq_with_pos'],
                'complementary_sites': result['complementary_sites'],
                'potential_kozak': kozak_sequences,
                'in_frame': triplet_stats['in_frame'],
                'codon_count': triplet_stats['codon_count'],
                'remaining_bases': triplet_stats['remaining_bases']
            }
            results.append(row_data)
            
            # Add row to HTML
            html_content += f"""
            <tr>
                <td>{row_data['transcript_id']}</td>
                <td class="sequence">{row_data['formatted_seq']}</td>
                <td class="sequence">{row_data['formatted_seq_with_pos']}</td>
                <td class="sequence">{row_data['formatted_pattern']}</td>
                <td class="sequence">{row_data['complementary_sites']}</td>
                <td class="sequence">{row_data['potential_kozak']}</td>
                <td>{row_data['in_frame']}</td>
                <td>{row_data['codon_count']}</td>
                <td>{row_data['remaining_bases']}</td>
            </tr>
            """
            
        html_content += """
            </table>
        </body>
        </html>
        """
        
        # Save main report
        output_file = os.path.join(output_dir, f"{self.prefix}_formatted_sequences.html")
        with open(output_file, 'w') as f:
            f.write(html_content)
            
        # Generate statistics report
        stats_html = f"""
        <html>
        <head>
            <style>
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ padding: 8px; text-align: left; border: 1px solid #ddd; }}
                th {{ background-color: #f2f2f2; }}
                .sequence {{ font-family: monospace; }}
            </style>
        </head>
        <body>
            <h2>Formatted Sequence Pattern Statistics</h2>
            <table>
                <tr>
                    <th>Formatted Pattern</th>
                    <th>Count</th>
                    <th>Count with Potential Complementary Sites</th>
                </tr>
        """
        
        # Store pattern stats for Excel export
        pattern_results = []
        
        # Sort patterns by count
        sorted_patterns = sorted(pattern_stats.items(), 
                               key=lambda x: x[1]['total'], 
                               reverse=True)
        
        for pattern, stats in sorted_patterns:
            total = stats['total']
            with_comp = stats['with_complementary']
            percentage = (with_comp / total * 100) if total > 0 else 0
            
            pattern_results.append({
                'pattern': pattern,
                'count': total,
                'with_complementary': f"{with_comp} ({percentage:.1f}%)"
            })
            
            stats_html += f"""
            <tr>
                <td class="sequence">{pattern}</td>
                <td>{total}</td>
                <td>{with_comp} ({percentage:.1f}%)</td>
            </tr>
            """
        
        stats_html += """
            </table>
        </body>
        </html>
        """
        
        # Save statistics report
        stats_file = os.path.join(output_dir, f"{self.prefix}_formatted_sequences_stats.html")
        with open(stats_file, 'w') as f:
            f.write(stats_html)
        
        # Export to Excel
        self.export_to_excel(results, pattern_results, output_dir)
        
        print(f"Saved formatted sequences report to {output_file}")
        print(f"Saved pattern statistics to {stats_file}")

    def export_to_excel_(self, sequence_results, pattern_results, output_dir):
        """
        Export results to Excel files.
        """
        # Export sequence results
        seq_df = pd.DataFrame(sequence_results)
        seq_excel_file = os.path.join(output_dir, f"{self.prefix}_formatted_sequences.xlsx")
        seq_df.to_excel(seq_excel_file, index=False)
        
        # Export pattern statistics
        stats_df = pd.DataFrame(pattern_results)
        stats_excel_file = os.path.join(output_dir, f"{self.prefix}_formatted_sequences_stats.xlsx")
        stats_df.to_excel(stats_excel_file, index=False)
        
        print(f"Saved sequence Excel report to {seq_excel_file}")
        print(f"Saved statistics Excel report to {stats_excel_file}")

    def export_to_excel(self, sequence_results, pattern_results, output_dir):
        def preprocess_value(value):
            if isinstance(value, list) and len(value) == 0:
                return ''
            return value

        def apply_color_format(cell, text):
            # 定义颜色对照表
            color_map = {
                'red': 'FF0000',
                'blue': '0000FF',
                'gray': '808080',
                'yellow': 'FFFF00',
            }
            
            # 如果text是列表,则将其连接成字符串
            if isinstance(text, list):
                text = ' '.join(text)
            
            # 如果text是None,则将其转换为空字符串
            if text is None:
                text = ''
            
            segments = text.split('<span style="color:')
            current_position = 0
            
            for segment in segments:
                if '">' in segment:
                    color, content = segment.split('">', 1)
                    content = content.replace('</span>', '')
                    
                    if current_position == 0:
                        cell.value = content
                    else:
                        cell.value += content
                    
                    start = current_position
                    end = current_position + len(content)
                    
                    font_color = color_map.get(color, '000000')
                    cell.font = Font(color=font_color)
                    
                    current_position = end
                else:
                    cell.value += segment
            
            return cell

        # 创建formatted_sequences.xlsx文件
        seq_workbook = Workbook()
        seq_sheet = seq_workbook.active
        seq_sheet.title = "Formatted Sequences"
        
        # 写入表头
        header = ['Transcript ID', 'Formatted Sequence', 'Formatted Sequence with Position', 
                  'Formatted Pattern', 'Potential Complementary Sites', 'Potential Kozak', 
                  'In Frame?', 'Codon Count', 'Remaining Bases']
        seq_sheet.append(header)
        
        # 写入数据
        for row in sequence_results:
            row_data = [
                preprocess_value(row['transcript_id']),
                preprocess_value(row['formatted_seq']),
                preprocess_value(row['formatted_seq_with_pos']),
                preprocess_value(row['formatted_pattern']),
                preprocess_value(row['complementary_sites']),
                preprocess_value(row['potential_kozak']),
                preprocess_value(row['in_frame']),
                preprocess_value(row['codon_count']),
                preprocess_value(row['remaining_bases'])
            ]
            seq_sheet.append(row_data)
            
            for col, value in enumerate(row_data[1:6], start=2):
                apply_color_format(seq_sheet.cell(row=seq_sheet.max_row, column=col), value)
        
        # 保存formatted_sequences.xlsx文件
        seq_excel_file = os.path.join(output_dir, f"{self.prefix}_formatted_sequences.xlsx")
        seq_workbook.save(seq_excel_file)
        print(f"Saved sequence Excel report to {seq_excel_file}")

        # 创建formatted_sequences_stats.xlsx文件  
        stats_workbook = Workbook()
        stats_sheet = stats_workbook.active
        stats_sheet.title = "Pattern Statistics"
        
        # 写入表头
        stats_header = ['Pattern', 'Count', 'Count with Potential Complementary Sites']
        stats_sheet.append(stats_header)
        
        # 写入数据
        for row in pattern_results:
            row_data = [
                preprocess_value(row['pattern']),
                preprocess_value(row['count']),
                preprocess_value(row['with_complementary'])
            ]
            stats_sheet.append(row_data)
            
            apply_color_format(stats_sheet.cell(row=stats_sheet.max_row, column=1), row_data[0])
        
        # 保存formatted_sequences_stats.xlsx文件
        stats_excel_file = os.path.join(output_dir, f"{self.prefix}_formatted_sequences_stats.xlsx")
        stats_workbook.save(stats_excel_file)
        print(f"Saved statistics Excel report to {stats_excel_file}")
    
    def format_sequence_with_predictions(self, transcript_id):
        """
        Enhanced format_sequence_with_predictions method with new features.
        """
        if transcript_id not in self.data or transcript_id not in self.sequences:
            return {
                'transcript_id': transcript_id,
                'formatted_seq': "",
                'formatted_pattern': "",
                'formatted_seq_with_pos': "",
                'potential_complementary_sites': "",
                'potential_kozak': "",
                'triplet_stats': {'in_frame': 'N/A', 'codon_count': 0, 'remaining_bases': 0}
            }
             
        predictions = self.data[transcript_id]['predictions']
        sequence = self.sequences[transcript_id]
        formatted_parts = []
        formatted_parts_with_pos = []
        complementary_sites_parts = []
        
        # Store information about found sites
        found_starts = []
        found_stops = []
        
        i = 0
        while i < len(predictions):
            current_pred = predictions[i]
            
            if current_pred in [0, 1]:
                start = i
                while i < len(predictions) and predictions[i] == current_pred:
                    i += 1
                
                # Add context bases for incomplete codons
                if (i - start) % 3 != 0:
                    if start > 0 and predictions[start-1] == 2:
                        context_base = sequence[start-1]
                        formatted_parts.append(
                            f'<span style="color:gray">{context_base.lower()}</span>'
                        )
                    if i < len(predictions) and predictions[i] == 2:
                        context_base = sequence[i]
                        formatted_parts.append(
                            f'<span style="color:gray">{context_base.lower()}</span>'
                        )
                
                color = 'red' if current_pred == 0 else 'blue'
                bases = sequence[start:i]
                
                if current_pred == 0:
                    found_starts.append((start, bases))
                else:
                    found_stops.append((start, bases))
                
                formatted_parts.append(
                    f'<span style="color:{color}">{bases.upper()}</span>'
                )
                formatted_parts_with_pos.append(
                    f'<span style="color:{color}">{bases.upper()}<sub>{start+1}</sub></span>'
                )
            else:
                start = i
                while i < len(predictions) and predictions[i] == 2:
                    i += 1
                
                gap_length = i - start
                if gap_length < 4:
                    bases = sequence[start:i]
                    formatted_str = f'<span style="color:gray">{bases.lower()}</span>'
                    formatted_parts.append(formatted_str)
                    formatted_parts_with_pos.append(formatted_str)
                else:
                    formatted_str = f'<span style="color:gray">—</span>'
                    formatted_parts.append(formatted_str)
                    formatted_parts_with_pos.append(formatted_str)
        # Generate complementary sites content
        if len(found_starts) == 1 and len(found_stops) == 0:
            # Only TIS found, look for potential stop codons
            potential_sites = self.find_complementary_sites(sequence, 'start', found_starts[0][0])
            if potential_sites:
                complementary_parts = []
                for pos, codon in potential_sites:
                    complementary_parts.append(
                        f'<span style="color:purple">{codon}<sub>{pos+1}</sub></span>'
                    )
                complementary_sites_parts = '<span style="color:gray">—</span>'.join(complementary_parts)
                
        elif len(found_stops) == 1 and len(found_starts) == 0:
            # Only TTS found, look for potential start codons
            potential_sites = self.find_complementary_sites(sequence, 'stop', found_stops[0][0])
            if potential_sites:
                complementary_parts = []
                for pos, codon in potential_sites:
                    complementary_parts.append(
                        f'<span style="color:purple">{codon}<sub>{pos+1}</sub></span>'
                    )
                complementary_sites_parts = '<span style="color:gray">—</span>'.join(complementary_parts)
        
        return {
            'transcript_id': transcript_id,
            'formatted_seq': ''.join(formatted_parts),
            'formatted_seq_with_pos': ''.join(formatted_parts_with_pos),
            'formatted_pattern': self.format_pattern(predictions, sequence) or '<span style="color:gray">—</span>',
            'complementary_sites': complementary_sites_parts,
            'potential_kozak': self.find_kozak_sequences(sequence),
            'triplet_stats': self.find_triplet_stats(
                found_starts[0][0] if found_starts else None,
                found_stops[0][0] if found_stops else None
            )
        }

def main(args):
    """主函数"""
    # 创建输出目录
    os.makedirs(args.output_dir, exist_ok=True)
    
    # 初始化分析器
    analyzer = TranscriptAnalyzer(args.predictions_file, args.fasta_file, args.prefix)
    
    # 生成可视化
    analyzer.plot_transcript_heatmap(args.output_dir)
    analyzer.plot_density_distribution(args.output_dir)
    analyzer.plot_continuity_stats(args.output_dir)
    analyzer.generate_formatted_sequences_report(args.output_dir)
    #analyzer.generate_formatted_sequence_stats(args.output_dir)
    
    # 生成统计报告
    stats = analyzer.generate_statistics()
    
    # 保存统计结果
    report_file = os.path.join(args.output_dir, f"{args.prefix}_statistics_report.txt")
    with open(report_file, 'w') as f:
        f.write(f"Prediction Analysis Report - {args.prefix}\n")
        f.write("=====================================\n\n")
        
        f.write("General Statistics:\n")
        f.write("-----------------\n")
        f.write(f"Total transcripts analyzed: {stats['total_transcripts']}\n")
        f.write(f"Perfectly matching transcripts: {stats['matching_transcripts']} "
                f"({stats['matching_transcripts']/stats['total_transcripts']*100:.2f}%)\n\n")
        
        f.write("Transcript Type Distribution:\n")
        f.write("--------------------------\n")
        for t_type, count in stats['transcript_types'].items():
            f.write(f"{t_type}: {count} ({count/stats['total_transcripts']*100:.2f}%)\n")
        f.write("\n")
        
        f.write("TIS Statistics:\n")
        f.write("--------------\n")
        f.write(f"Total regions: {stats['tis_stats']['total_regions']}\n")
        f.write(f"Average length: {stats['tis_stats']['avg_length']:.2f} bp\n")
        f.write(f"Maximum length: {stats['tis_stats']['max_length']} bp\n\n")
        
        f.write("TTS Statistics:\n")
        f.write("--------------\n")
        f.write(f"Total regions: {stats['tts_stats']['total_regions']}\n")
        f.write(f"Average length: {stats['tts_stats']['avg_length']:.2f} bp\n")
        f.write(f"Maximum length: {stats['tts_stats']['max_length']} bp\n")
        
    print(f"Saved statistics report to {report_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Analyze prediction results')
    parser.add_argument('--predictions_file', type=str, required=True,
                      help='Path to the predictions pkl file')
    parser.add_argument('--output_dir', type=str, required=True,
                      help='Directory to save analysis results')
    parser.add_argument('--prefix', type=str, required=True,
                       help='Prefix for output files')
    parser.add_argument('--fasta_file', type=str, required=True,
                       help='Path to FASTA file with transcript sequences')
    
    args = parser.parse_args()
    main(args)