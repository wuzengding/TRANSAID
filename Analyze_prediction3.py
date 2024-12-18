import argparse
import pickle
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
import seaborn as sns
from collections import defaultdict
import os
import random
from Bio import SeqIO
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class TranscriptAnalyzer:
    def __init__(self, predictions_file, fasta_file, gbff_file, prefix):
        """初始化分析器"""
        # 读取预测结果，包括概率值
        with open(predictions_file, 'rb') as f:
            raw_data = pickle.load(f)
            self.data = {item['transcript_id']: {
                'predictions': item['predictions'],
                'predictions_probs': item['predictions_probs'],  # 读取概率值
                'true_labels': item['true_labels'],
                'length': item['length'],
                'is_match': item['is_match']
            } for item in raw_data}
            
        # 读取序列数据
        self.sequences = {}
        for record in SeqIO.parse(fasta_file, "fasta"):
            transcript_id = record.id.split('.')[0]
            self.sequences[transcript_id] = str(record.seq)
        
        # 读取 CDS 位置信息
        self.true_positions = {}
        for record in SeqIO.parse(gbff_file, "genbank"):
            trans_id = record.id.split(".")[0]
            for feature in record.features:
                if feature.type == "CDS":
                    start = int(feature.location.start)
                    end = int(feature.location.end)
                    tts = end - 2  # Get first base of stop codon
                    self.true_positions[trans_id] = (start, tts)

        # 检查ID匹配情况
        missing_ids = set(self.data.keys()) - set(self.true_positions.keys())
        if missing_ids:
            print(f"Warning: {len(missing_ids)} transcripts have no CDS annotation:")
            for mid in list(missing_ids)[:5]:  # 只打印前5个作为示例
                print(f"  - {mid}")

        self.formatted_sequence_stats = defaultdict(int)
        self.prefix = prefix
    
    def normalize_transcript(self, predictions, length):
        """将转录本标准化到0-1区间"""
        norm_positions = np.linspace(0, 1, length)
        return norm_positions, predictions
        
    def find_continuous_regions(self, predictions, label_type):
        """找到连续的预测区域"""
        regions = []
        current_start = None
        current_length = 0
        
        for i, pred in enumerate(predictions):
            if pred == label_type:
                if current_start is None:
                    current_start = i
                current_length += 1
            else:
                if current_length > 0:
                    regions.append({
                        'start': current_start,
                        'length': current_length,
                        'rel_start': current_start / len(predictions)
                    })
                current_start = None
                current_length = 0
                
        if current_length > 0:
            regions.append({
                'start': current_start,
                'length': current_length,
                'rel_start': current_start / len(predictions)
            })
            
        return regions

    def analyze_transcript(self, transcript_id):
        """分析单个转录本"""
        transcript_data = self.data[transcript_id]
        predictions = transcript_data['predictions']
        length = transcript_data['length']
        
        # 找到TIS和TTS区域
        tis_regions = self.find_continuous_regions(predictions, 0)
        tts_regions = self.find_continuous_regions(predictions, 1)
        
        return {
            'tis_regions': tis_regions,
            'tts_regions': tts_regions,
            'normalized_predictions': self.normalize_transcript(predictions, length)
        }

    def plot_transcript_heatmap(self, output_dir, group_size=400):
        """
        绘制按长度分组的转录本预测热图,每组固定group_size个转录本
        """
        # 获取所有转录本长度并排序
        transcript_lengths = [(tid, d['length']) for tid, d in self.data.items()]
        transcript_lengths.sort(key=lambda x: x[1])
        
        # 将转录本分组
        num_groups = (len(transcript_lengths) + group_size - 1) // group_size
        groups = [transcript_lengths[i*group_size : (i+1)*group_size] 
                 for i in range(num_groups)]
        
        # 创建颜色映射 (TIS=红色, TTS=蓝色, non-TIS/TTS=浅灰色, 背景=白色)
        colors = ['red', 'blue', '#EEEEEE', 'white']
        cmap = ListedColormap(colors)
        
        # 为每个组绘制热图
        for idx, group in enumerate(groups):
            group_min = group[0][1]
            group_max = group[-1][1]
            
            # 创建数据矩阵
            data_matrix = np.full((len(group), group_max), 3)  # 背景为白色
            
            # 填充预测数据
            for row, (tid, length) in enumerate(group):
                predictions = self.data[tid]['predictions']
                data_matrix[row, :length] = predictions
            
            # 绘制热图
            plt.figure(figsize=(15, 10))
            
            # 绘制热图并保存返回的图像对象
            im = plt.imshow(data_matrix, aspect='auto', cmap=cmap, interpolation='none')
            
            # 设置标题和标签
            plt.title(f'{self.prefix} Length {group_min}-{group_max}bp (n={len(group)})')
            plt.ylabel('Transcripts')
            plt.xlabel('Position (bp)')

            # 修改后的 colorbar 创建方式
            cbar = plt.colorbar(im, ticks=[0, 1, 2])
            cbar.ax.set_yticklabels(['TIS', 'TTS', 'non-TIS/TTS'])  # 设置标签
            plt.yticks([])
            
            # 保存图片
            output_file = os.path.join(output_dir, 
                         f"{self.prefix}_heatmap_length{group_min}-{group_max}.png")
            plt.savefig(output_file, dpi=300, bbox_inches='tight')
            plt.close()
            print(f"Saved heatmap to {output_file}")

    def plot_density_distribution(self, output_dir):
        """绘制TIS/TTS位置分布密度图"""
        plt.figure(figsize=(12, 6))
        
        # 收集TIS和TTS位置
        tis_positions = []
        tts_positions = []
        
        for transcript_id in self.data:
            analysis = self.analyze_transcript(transcript_id)
            tis_positions.extend([r['rel_start'] for r in analysis['tis_regions']])
            tts_positions.extend([r['rel_start'] for r in analysis['tts_regions']])
        
        # 绘制密度图
        sns.kdeplot(data=tis_positions, color='red', label='TIS')
        sns.kdeplot(data=tts_positions, color='blue', label='TTS')
        
        plt.xlabel('Relative Position in Transcript')
        plt.ylabel('Density')
        plt.title(f'{self.prefix} TIS/TTS Position Distribution')
        plt.legend()
        
        output_file = os.path.join(output_dir, f"{self.prefix}_position_density.png")
        plt.savefig(output_file)
        plt.close()
        print(f"Saved density plot to {output_file}")

    def plot_continuity_stats(self, output_dir):
        """绘制连续区域长度分布"""
        tis_lengths = []
        tts_lengths = []
        
        for transcript_id in self.data:
            analysis = self.analyze_transcript(transcript_id)
            tis_lengths.extend([r['length'] for r in analysis['tis_regions']])
            tts_lengths.extend([r['length'] for r in analysis['tts_regions']])
        
        plt.figure(figsize=(12, 6))
        plt.hist([tis_lengths, tts_lengths], 
                label=['TIS', 'TTS'], 
                bins=30, 
                alpha=0.7)
        plt.xlabel('Continuous Region Length (bp)')
        plt.ylabel('Frequency')
        plt.title(f'{self.prefix} Distribution of Continuous TIS/TTS Regions')
        plt.legend()
        
        output_file = os.path.join(output_dir, f"{self.prefix}_continuity_distribution.png")
        plt.savefig(output_file)
        plt.close()
        print(f"Saved continuity distribution to {output_file}")

    def generate_statistics(self):
        """生成统计摘要"""
        stats = {
            'total_transcripts': len(self.data),
            'matching_transcripts': sum(1 for d in self.data.values() if d['is_match']),
            'tis_stats': defaultdict(float),
            'tts_stats': defaultdict(float),
            'transcript_types': defaultdict(int)
        }
        
        # 统计每种转录本类型的数量
        for tid in self.data:
            trans_type = tid.split('_')[0]
            stats['transcript_types'][trans_type] += 1
        
        # TIS和TTS统计
        for transcript_id in self.data:
            analysis = self.analyze_transcript(transcript_id)
            
            # TIS统计
            tis_regions = analysis['tis_regions']
            stats['tis_stats']['total_regions'] += len(tis_regions)
            if tis_regions:
                stats['tis_stats']['avg_length'] += sum(r['length'] for r in tis_regions)
                stats['tis_stats']['max_length'] = max(
                    stats['tis_stats']['max_length'],
                    max(r['length'] for r in tis_regions)
                )
            
            # TTS统计
            tts_regions = analysis['tts_regions']
            stats['tts_stats']['total_regions'] += len(tts_regions)
            if tts_regions:
                stats['tts_stats']['avg_length'] += sum(r['length'] for r in tts_regions)
                stats['tts_stats']['max_length'] = max(
                    stats['tts_stats']['max_length'],
                    max(r['length'] for r in tts_regions)
                )
        
        # 计算平均值
        if stats['tis_stats']['total_regions'] > 0:
            stats['tis_stats']['avg_length'] /= stats['tis_stats']['total_regions']
        if stats['tts_stats']['total_regions'] > 0:
            stats['tts_stats']['avg_length'] /= stats['tts_stats']['total_regions']
            
        return stats
    
    def collect_pattern_statistics(self):
        """
        收集所有格式化序列的统计信息
        """
        pattern_stats = defaultdict(int)
        total_count = 0
        
        for transcript_id in self.data.keys():
            result = self.format_sequence_with_predictions(transcript_id)
            pattern = result['formatted_seq']
            pattern_stats[pattern] += 1
            total_count += 1
        
        # 转换为排序后的列表
        sorted_stats = [(pattern, count, count/total_count*100) 
                       for pattern, count in pattern_stats.items()]
        sorted_stats.sort(key=lambda x: x[1], reverse=True)
        
        return sorted_stats, total_count


    def find_complementary_sites(self, sequence, current_site_type, current_pos):
        """
        Find potential complementary translation sites.
        
        Args:
            sequence (str): The complete transcript sequence
            current_site_type (str): 'start' or 'stop' indicating what we're looking for
            current_pos (int): Position of the current site
            
        Returns:
            list: List of tuples containing (position, codon)
        """
        complementary_sites = []
        seq_len = len(sequence)
        
        if current_site_type == 'start':
            # Looking for stop codons downstream
            pos = current_pos + 3  # Start after ATG
            while pos + 2 < seq_len:
                codon = sequence[pos:pos+3].upper()
                if codon in ['TAA', 'TAG', 'TGA']:
                    complementary_sites.append((pos, codon))
                pos += 3  # Move to next codon
        else:
            # Looking for start codons upstream
            pos = current_pos - 3  # Start before stop codon
            while pos >= 0:
                codon = sequence[pos:pos+3].upper()
                if codon == 'ATG':
                    complementary_sites.append((pos, codon))
                pos -= 3  # Move to previous codon
                
        return complementary_sites
        
    def find_triplet_stats(self, start_pos, stop_pos):
        """
        Calculate triplet codon statistics between start and stop positions.
        
        Args:
            start_pos (int or None): Position of start codon
            stop_pos (int or None): Position of stop codon
            
        Returns:
            dict: Dictionary containing in_frame status, codon count, and remaining bases
        """
        # Handle cases where either position is None
        if start_pos is None or stop_pos is None:
            return {
                'in_frame': 'N/A',
                'codon_count': 0,
                'remaining_bases': 0
            }
            
        # Ensure stop_pos is greater than start_pos
        if stop_pos <= start_pos:
            return {
                'in_frame': 'N/A',
                'codon_count': 0,
                'remaining_bases': 0
            }
        
        distance = stop_pos - start_pos
        codon_count = distance // 3
        remaining_bases = distance % 3
        is_triplet = remaining_bases == 0
        
        return {
            'in_frame': 'Yes' if is_triplet else 'No',
            'codon_count': codon_count,
            'remaining_bases': remaining_bases
        }

    def evaluate_kozak_sequence(self, sequence, atg_pos):
        """
        Evaluate Kozak sequence strength based on position weight matrix.
        
        Args:
            sequence (str): Full transcript sequence
            atg_pos (int): Position of the ATG start codon
            
        Returns:
            tuple: (kozak_sequence, strength)
        """
        if atg_pos < 6 or atg_pos + 4 >= len(sequence):
            return None, None
            
        # Extract Kozak context (-6 to +4)
        kozak = sequence[atg_pos-6:atg_pos+4].upper()
        if len(kozak) != 10:
            return None, None
            
        # Check core positions
        minus_3 = kozak[3]  # -3 position
        plus_4 = kozak[9]   # +4 position
        
        # Determine strength
        if minus_3 in ['A', 'G'] and plus_4 == 'G':
            strength = 'S'  # Strong
        elif minus_3 in ['A', 'G'] or plus_4 == 'G':
            strength = 'M'  # Moderate
        else:
            return None, None  # Don't report weak sequences
            
        return kozak, strength

    def find_kozak_sequences(self, sequence):
        """
        Find all valid Kozak sequences in the transcript.
        
        Args:
            sequence (str): Full transcript sequence
            
        Returns:
            list: List of formatted Kozak sequences with positions and strengths
        """
        kozak_sites = []
        pos = 0
        
        while True:
            pos = sequence.upper().find('ATG', pos)
            if pos == -1:
                break
                
            kozak_seq, strength = self.evaluate_kozak_sequence(sequence, pos)
            if kozak_seq and strength:
                # Format: first 6 bases + ATG with position + last base + strength
                formatted_kozak = (
                    f"{kozak_seq[:6].lower()}"
                    f"ATG<sub>{pos+1}</sub>"
                    f"{kozak_seq[9].lower()}"
                    f"({strength})"
                )
                kozak_sites.append(formatted_kozak)
            
            pos += 3
            
        return '; '.join(kozak_sites) if kozak_sites else ""

    def format_pattern(self, predictions, sequence):
        """
        Create simplified pattern showing only TIS and TTS positions.
    
        Args:
            predictions (np.array): Prediction array
            sequence (str): Original sequence
        
        Returns:
            str: Formatted pattern string
        """
        pattern_parts = []
        i = 0
        current_type = None
        current_bases = []
    
        while i < len(predictions):
            if predictions[i] in [0, 1]:
                if current_type is None:
                    current_type = predictions[i]
                    current_bases = [sequence[i].upper()]
                elif current_type == predictions[i]:
                    current_bases.append(sequence[i].upper())
                else:
                    # Add current group
                    color = 'red' if current_type == 0 else 'blue'
                    pattern_parts.append(
                        f'<span style="color:{color}">{"".join(current_bases)}</span>'
                    )
                    current_type = predictions[i]
                    current_bases = [sequence[i].upper()]
                i += 1
            else:
                if current_bases:
                    color = 'red' if current_type == 0 else 'blue'
                    pattern_parts.append(
                        f'<span style="color:{color}">{"".join(current_bases)}</span>'
                    )
                    current_type = None
                    current_bases = []
                if len(pattern_parts) > 0 and pattern_parts[-1] != '<span style="color:gray">—</span>':
                    pattern_parts.append('<span style="color:gray">—</span>')
                while i < len(predictions) and predictions[i] not in [0, 1]:
                    i += 1
        
        if current_bases:
            color = 'red' if current_type == 0 else 'blue'
            pattern_parts.append(
                f'<span style="color:{color}">{"".join(current_bases)}</span>'
            )
        
        return ''.join(pattern_parts) if pattern_parts else '<span style="color:gray">—</span>'

    def strip_html_tags(self, html_text):
        """
        Remove HTML tags while preserving the text content
    
        Args:
            html_text (str): HTML formatted text
        
        Returns:
            str: Plain text without HTML tags
        """
        if not html_text:
            return ""
        
        result = ""
        i = 0
        while i < len(html_text):
            if html_text[i:i+5] == '<span':
                # Skip span tag
                i = html_text.find('>', i) + 1
            elif html_text[i:i+7] == '</span>':
                # Skip closing span tag
                i += 7
            elif html_text[i:i+5] == '<sub>':
                # Skip sub tag
                i = html_text.find('</sub>', i) + 6
            else:
                result += html_text[i]
                i += 1
            
        return result
        
    def export_to_excel(self, sequence_results, output_dir):
        """
        Export results to Excel files with proper formatting.
        """
        # Create workbooks
        seq_workbook = Workbook()
        seq_sheet = seq_workbook.active
        seq_sheet.title = "Formatted Sequences"
    
        # Write sequence data headers 
        headers = [
            'Transcript_ID',
            'Formatted_Sequence',
            'Formatted_Sequence_with_Position',
            'Formatted_Pattern',
            'Formatted_Pattern_score',
            'Potential_Complementary_Sites',
            'Potential_Kozak',
            'True_TIS',    # 移到正确位置
            'True_TTS',    # 移到正确位置
            'In_Frame?',
            'Codon_Count',
            'Remaining_Bases'
            ]
        seq_sheet.append(headers)
    
        # Write sequence data
        for row in sequence_results:
            # Get true positions
            trans_id = row['transcript_id']
            true_tis, true_tts = self.true_positions.get(trans_id, ('—', '—'))
        
            row_data = [
                row['transcript_id'],
                self.format_text_for_excel(row['formatted_seq']),
                self.format_text_for_excel(row['formatted_seq_with_pos']),
                self.format_text_for_excel(row['formatted_pattern']),
                row['formatted_probs'], 
                self.format_text_for_excel(row['complementary_sites']),
                self.format_text_for_excel(row['potential_kozak']),
                true_tis, 
                true_tts, 
                row['in_frame'],
                row['codon_count'],
                row['remaining_bases']
            ]
            seq_sheet.append(row_data)
        
            # Apply color formatting
            self.apply_excel_formatting(seq_sheet.cell(row=seq_sheet.max_row, column=2), row['formatted_seq'])
            self.apply_excel_formatting(seq_sheet.cell(row=seq_sheet.max_row, column=3), row['formatted_seq_with_pos'])
            self.apply_excel_formatting(seq_sheet.cell(row=seq_sheet.max_row, column=4), row['formatted_pattern'])
            self.apply_excel_formatting(seq_sheet.cell(row=seq_sheet.max_row, column=5), row['formatted_probs'])
            self.apply_excel_formatting(seq_sheet.cell(row=seq_sheet.max_row, column=6), row['complementary_sites'])
            self.apply_excel_formatting(seq_sheet.cell(row=seq_sheet.max_row, column=7), row['potential_kozak']) 
        
        # Save sequence Excel file
        seq_excel_file = os.path.join(output_dir, f"{self.prefix}_formatted_sequences.xlsx")
        seq_workbook.save(seq_excel_file)

    def format_text_for_excel(self, html_text, keep_position=False):
        """
        Format HTML text for Excel while optionally preserving position numbers
    
        Args:
            html_text (str): HTML formatted text
            keep_position (bool): Whether to keep position numbers
        
        Returns:
            str: Formatted text for Excel
        """
        if not html_text:
            return ""
    
        result = ""
        i = 0
        while i < len(html_text):
            if html_text[i:i+5] == '<span':
                # Skip span tag
                i = html_text.find('>', i) + 1
            elif html_text[i:i+7] == '</span>':
                # Skip closing span tag
                i += 7
            elif html_text[i:i+5] == '<sub>':
                if keep_position:
                    # Keep position number with underscore prefix
                    sub_end = html_text.find('</sub>', i)
                    position = html_text[i+5:sub_end]
                    result += position
                i = html_text.find('</sub>', i) + 6
            else:
                result += html_text[i]
                i += 1

        return result

    def apply_excel_formatting(self, cell, html_content):
        """
        Apply color formatting to Excel cell based on HTML content
        """
        # Process the text
        text = []
        current_text = ''
        is_dash = False
        i = 0
    
        while i < len(html_content):
            if html_content[i:i+6] == '<span ':
                # If we have accumulated text, save it
                if current_text:
                    text.append(current_text)
                    current_text = ''
            
                # Check if this is a dash
                if '—' in html_content[i:html_content.find('</span>', i)]:
                    is_dash = True
            
                # Skip to content
                i = html_content.find('">', i) + 2
                continue
            elif html_content[i:i+7] == '</span>':
                if current_text:
                    text.append(current_text)
                    current_text = ''
                i += 7
                continue
            elif html_content[i:i+5] == '<sub>':
                # Handle position subscript
                sub_end = html_content.find('</sub>', i)
                if sub_end != -1:
                    position = html_content[i+5:sub_end]
                    current_text += position  # 直接添加位置数字
                    i = sub_end + 6
                    continue
        
            current_text += html_content[i]
            i += 1
    
        # Add any remaining text
        if current_text:
            text.append(current_text)
    
        # Combine all text segments
        cell.value = ''.join(text)
    
        # Only apply gray color for dash
        if is_dash:
            cell.font = Font(color='808080')  # 灰色
            
    def generate_formatted_sequences_report(self, output_dir):
        """Generate HTML report with true TIS/TTS positions"""
        html_content = """
        <html>
        <head>
            <style>
                table { border-collapse: collapse; width: 100%; margin-top: 20px; }
                th, td { padding: 8px; text-align: left; border: 1px solid #ddd; }
                th { background-color: #f2f2f2; }
                td { vertical-align: top; }
                .sequence { font-family: monospace; white-space: pre-wrap; }
                .group-header { background-color: #e6e6e6; }
            </style>
        </head>
        <body>
            <h2>Formatted Sequence Report</h2>
            <table>
                <tr>
                    <th rowspan="2">Transcript_ID</th>
                    <th rowspan="2">Formatted_Sequence</th>
                    <th rowspan="2">Formatted_Sequence with Position</th>
                    <th rowspan="2">Formatted_Pattern</th>
                    <th rowspan="2">Formatted_Pattern_score</th> 
                    <th rowspan="2">Potential_Complementary_Sites</th>
                    <th rowspan="2">Potential_Kozak</th>
                    <th rowspan="2">True_TIS</th>
                    <th rowspan="2">True_TTS</th>
                    <th colspan="3" class="group-header">Right Triplet?</th>
                </tr>
                <tr>
                    <th>In Frame?</th>
                    <th>Codon Count</th>
                    <th>Remaining Bases</th>
                </tr>
        """
        
        # Process each transcript
        results = []
        for transcript_id in sorted(self.data.keys()):
            # Get formatted data
            result = self.format_sequence_with_predictions(transcript_id)
            
            # Get true positions
            true_tis, true_tts = self.true_positions.get(transcript_id, ('——', '——'))
            
            # Add to HTML
            html_content += f"""
            <tr>
                <td>{transcript_id}</td>
                <td class="sequence">{result['formatted_seq']}</td>
                <td class="sequence">{result['formatted_seq_with_pos']}</td>
                <td class="sequence">{result['formatted_pattern']}</td>
                <td class="sequence">{result['formatted_probs']}</td>
                <td class="sequence">{result['complementary_sites']}</td>
                <td class="sequence">{result['potential_kozak']}</td>
                <td>{true_tis}</td>
                <td>{true_tts}</td>
                <td>{result['triplet_stats']['in_frame']}</td>
                <td>{result['triplet_stats']['codon_count']}</td>
                <td>{result['triplet_stats']['remaining_bases']}</td>
            </tr>
            """
            
            results.append({
                'transcript_id': transcript_id,
                'formatted_seq': result['formatted_seq'],
                'formatted_seq_with_pos': result['formatted_seq_with_pos'],
                'formatted_pattern': result['formatted_pattern'],
                'formatted_probs': result['formatted_probs'],  
                'true_tis': true_tis,
                'true_tts': true_tts,
                'complementary_sites': result['complementary_sites'],
                'potential_kozak': result['potential_kozak'],
                'in_frame': result['triplet_stats']['in_frame'],
                'codon_count': result['triplet_stats']['codon_count'],
                'remaining_bases': result['triplet_stats']['remaining_bases']
            })
        
        html_content += """
            </table>
        </body>
        </html>
        """
        
        # Save reports
        main_file = os.path.join(output_dir, f"{self.prefix}_formatted_sequences.html")
        with open(main_file, 'w') as f:
            f.write(html_content)
        
        return results

    def format_sequence_with_predictions(self, transcript_id):
        """
        Enhanced format_sequence_with_predictions method with new features.
        """
        if transcript_id not in self.data or transcript_id not in self.sequences:
            return {
                'transcript_id': transcript_id,
                'formatted_seq': "",
                'formatted_seq_with_pos': "",
                'formatted_pattern': "",
                'formatted_probs': "",
                'complementary_sites': "",
                'potential_kozak': "",
                'triplet_stats': {
                    'in_frame': 'N/A', 
                    'codon_count': 0, 
                    'remaining_bases': 0},
            }
             
        predictions = self.data[transcript_id]['predictions']
        predictions_probs = self.data[transcript_id]['predictions_probs']  # 添加概率值读取
        sequence = self.sequences[transcript_id]
        formatted_parts = []
        formatted_parts_with_pos = []
        complementary_sites_parts = []
        prob_parts = []  # 添加概率值格式化列表

        #print(f"transcript_id: {transcript_id}")
        #print(f"predictions length: {len(predictions)}")  
        #print(f"predictions_probs length: {len(predictions_probs)}")
        # Store information about found sites
        found_starts = []
        found_stops = []
    
        i = 0
        while i < len(predictions):
            current_pred = predictions[i]
        
            if current_pred in [0, 1]:
                start = i
                while i < len(predictions) and predictions[i] == current_pred:
                    # 添加概率值格式化
                    probs = predictions_probs[i]
                    prob_parts.append(f"[{probs[0]:.2f},{probs[1]:.2f},{probs[2]:.2f}]")
                    i += 1
            
                # Add context bases for incomplete codons
                if (i - start) % 3 != 0:
                    if start > 0 and predictions[start-1] == 2:
                        context_base = sequence[start-1]
                        formatted_parts.append(
                            f'<span style="color:gray">{context_base.lower()}</span>'
                        )
                    if i < len(predictions) and predictions[i] == 2:
                        context_base = sequence[i]
                        formatted_parts.append(
                            f'<span style="color:gray">{context_base.lower()}</span>'
                        )
            
                color = 'red' if current_pred == 0 else 'blue'
                bases = sequence[start:i]
            
                if current_pred == 0:
                    found_starts.append((start, bases))
                else:
                    found_stops.append((start, bases))
            
                formatted_parts.append(
                    f'<span style="color:{color}">{bases.upper()}</span>'
                )
                formatted_parts_with_pos.append(
                    f'<span style="color:{color}">{bases.upper()}<sub>{start+1}</sub></span>'
                )
            
                # 添加概率值分隔符
                if i < len(predictions):
                    prob_parts.append("")
                
            else:
                start = i
                while i < len(predictions) and predictions[i] == 2:
                    i += 1
            
                gap_length = i - start
                if gap_length < 4:
                    bases = sequence[start:i]
                    formatted_str = f'<span style="color:gray">{bases.lower()}</span>'
                    formatted_parts.append(formatted_str)
                    formatted_parts_with_pos.append(formatted_str)
                else:
                    formatted_str = f'<span style="color:gray">—</span>'
                    formatted_parts.append(formatted_str)
                    formatted_parts_with_pos.append(formatted_str)
            
                # 添加概率值横杠
                prob_parts.append("—")
                if i < len(predictions) and predictions[i] in [0, 1]:
                    prob_parts.append("")

        # Generate complementary sites content
        if len(found_starts) == 1 and len(found_stops) == 0:
            # Only TIS found, look for potential stop codons
            potential_sites = self.find_complementary_sites(sequence, 'start', found_starts[0][0])
            if potential_sites:
                complementary_parts = []
                for pos, codon in potential_sites:
                    complementary_parts.append(
                        f'<span style="color:purple">{codon}<sub>{pos+1}</sub></span>'
                    )
                complementary_sites_parts = '<span style="color:gray">—</span>'.join(complementary_parts)
                
        elif len(found_stops) == 1 and len(found_starts) == 0:
            # Only TTS found, look for potential start codons
            potential_sites = self.find_complementary_sites(sequence, 'stop', found_stops[0][0])
            if potential_sites:
                complementary_parts = []
                for pos, codon in potential_sites:
                    complementary_parts.append(
                        f'<span style="color:purple">{codon}<sub>{pos+1}</sub></span>'
                    )
                complementary_sites_parts = '<span style="color:gray">—</span>'.join(complementary_parts)
        
        return {
            'transcript_id': transcript_id,
            'formatted_seq': ''.join(formatted_parts),
            'formatted_seq_with_pos': ''.join(formatted_parts_with_pos),
            'formatted_pattern': self.format_pattern(predictions, sequence) or '<span style="color:gray">—</span>',
            'complementary_sites': complementary_sites_parts,
            'potential_kozak': self.find_kozak_sequences(sequence),
            'triplet_stats': self.find_triplet_stats(
                found_starts[0][0] if found_starts else None,
                found_stops[0][0] if found_stops else None
            ),
            'formatted_probs': ''.join(prob_parts)  # 添加格式化后的概率值
        }

def main(args):
    """主函数"""
    # 创建输出目录
    os.makedirs(args.output_dir, exist_ok=True)
    
    # 初始化分析器
    analyzer = TranscriptAnalyzer(args.predictions_file, args.fasta_file, 
                                args.gbff_file, args.prefix)
    
    # 生成可视化
    analyzer.plot_transcript_heatmap(args.output_dir)
    analyzer.plot_density_distribution(args.output_dir)
    analyzer.plot_continuity_stats(args.output_dir)
    results = analyzer.generate_formatted_sequences_report(args.output_dir)
    #pattern_results = analyzer.collect_pattern_statistics()
    analyzer.export_to_excel(results, args.output_dir)
    
    # 生成统计报告
    stats = analyzer.generate_statistics()
    
    # 保存统计结果
    report_file = os.path.join(args.output_dir, f"{args.prefix}_statistics_report.txt")
    with open(report_file, 'w') as f:
        f.write(f"Prediction Analysis Report - {args.prefix}\n")
        f.write("=====================================\n\n")
        
        f.write("General Statistics:\n")
        f.write("-----------------\n")
        f.write(f"Total transcripts analyzed: {stats['total_transcripts']}\n")
        f.write(f"Perfectly matching transcripts: {stats['matching_transcripts']} "
                f"({stats['matching_transcripts']/stats['total_transcripts']*100:.2f}%)\n\n")
        
        f.write("Transcript Type Distribution:\n")
        f.write("--------------------------\n")
        for t_type, count in stats['transcript_types'].items():
            f.write(f"{t_type}: {count} ({count/stats['total_transcripts']*100:.2f}%)\n")
        f.write("\n")
        
        f.write("TIS Statistics:\n")
        f.write("--------------\n")
        f.write(f"Total regions: {stats['tis_stats']['total_regions']}\n")
        f.write(f"Average length: {stats['tis_stats']['avg_length']:.2f} bp\n")
        f.write(f"Maximum length: {stats['tis_stats']['max_length']} bp\n\n")
        
        f.write("TTS Statistics:\n")
        f.write("--------------\n")
        f.write(f"Total regions: {stats['tts_stats']['total_regions']}\n")
        f.write(f"Average length: {stats['tts_stats']['avg_length']:.2f} bp\n")
        f.write(f"Maximum length: {stats['tts_stats']['max_length']} bp\n")
        
    print(f"Saved statistics report to {report_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Analyze prediction results')
    parser.add_argument('--predictions_file', type=str, required=True,
                      help='Path to the predictions pkl file')
    parser.add_argument('--gbff_file', type=str, required=True,
                       help='Path to GBFF file with CDS annotations')
    parser.add_argument('--output_dir', type=str, required=True,
                      help='Directory to save analysis results')
    parser.add_argument('--prefix', type=str, required=True,
                       help='Prefix for output files')
    parser.add_argument('--fasta_file', type=str, required=True,
                       help='Path to FASTA file with transcript sequences')
    
    args = parser.parse_args()
    main(args)